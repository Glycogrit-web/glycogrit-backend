"""
Excel Import Service for Physical Rewards
Imports tracking data from Excel files (AWB numbers, tracking URLs, courier names)
"""
import csv
import io
import logging
from datetime import datetime, timezone
from typing import Dict, List, Optional

from openpyxl import load_workbook
from sqlalchemy.orm import Session
from sqlalchemy.sql import func

from app.core.enums import RewardStatus
from app.models.user_reward import UserReward

logger = logging.getLogger(__name__)


class ExcelImportService:
    """Service for importing tracking data from Excel/CSV files"""

    def __init__(self, db: Session):
        """
        Initialize Excel import service

        Args:
            db: Database session
        """
        self.db = db

    def import_tracking_data(
        self,
        event_id: int,
        file_content: bytes,
        filename: str,
        admin_id: int
    ) -> Dict:
        """
        Import tracking data from Excel or CSV file

        Expected columns (flexible matching):
        - Order Reference OR Reward ID (to match reward)
        - Tracking ID (required)
        - Tracking URL (optional)
        - Courier Name (optional)

        Processing:
        1. Read Excel/CSV file
        2. Find reward by order_reference or reward_id
        3. Update tracking fields
        4. Change status to TRACKING_ORDER
        5. Set tracking_visible_to_user = True

        Args:
            event_id: Event ID to filter rewards
            file_content: File content as bytes
            filename: Original filename
            admin_id: Admin user ID who is importing

        Returns:
            {
                total_rows: int,
                successful: int,
                failed: int,
                not_found: int,
                errors: List[str]
            }

        Raises:
            ValueError: If file format invalid or required columns missing
        """
        logger.info(f"📥 Importing tracking data for event {event_id} from file: {filename}")

        # Determine file type and convert to CSV if needed
        if filename.lower().endswith(('.xlsx', '.xls')):
            csv_content = self._convert_xlsx_to_csv(file_content)
        elif filename.lower().endswith('.csv'):
            csv_content = file_content.decode('utf-8')
        else:
            raise ValueError("File must be CSV (.csv) or Excel (.xlsx/.xls)")

        # Parse CSV
        rows = list(csv.DictReader(io.StringIO(csv_content)))

        if not rows:
            raise ValueError("File is empty or has no data rows")

        # Identify columns (flexible matching)
        headers = list(rows[0].keys())
        column_map = self._identify_columns(headers)

        logger.info(f"   Column mapping: {column_map}")

        # Validate required columns
        if not column_map.get('tracking_id'):
            raise ValueError(
                f"Required column 'Tracking ID' not found. "
                f"Available columns: {', '.join(headers)}"
            )

        if not (column_map.get('order_reference') or column_map.get('reward_id')):
            raise ValueError(
                f"Required column 'Order Reference' or 'Reward ID' not found. "
                f"Available columns: {', '.join(headers)}"
            )

        # Process rows
        stats = {
            'total_rows': len(rows),
            'successful': 0,
            'failed': 0,
            'not_found': 0,
            'errors': []
        }

        for row_num, row in enumerate(rows, start=2):  # Start from 2 (header is row 1)
            try:
                result = self._process_tracking_row(
                    row=row,
                    row_num=row_num,
                    column_map=column_map,
                    event_id=event_id,
                    admin_id=admin_id
                )

                if result['success']:
                    stats['successful'] += 1
                elif result.get('not_found'):
                    stats['not_found'] += 1
                    stats['errors'].append(result['error'])
                else:
                    stats['failed'] += 1
                    stats['errors'].append(result['error'])

            except Exception as e:
                error_msg = f"Row {row_num}: Unexpected error - {str(e)}"
                logger.error(f"❌ {error_msg}")
                stats['failed'] += 1
                stats['errors'].append(error_msg)

        # Commit all changes
        self.db.commit()

        logger.info(
            f"✅ Import complete: {stats['successful']} successful, "
            f"{stats['failed']} failed, {stats['not_found']} not found"
        )

        return stats

    def _convert_xlsx_to_csv(self, excel_content: bytes) -> str:
        """
        Convert Excel file to CSV format

        Args:
            excel_content: Excel file as bytes

        Returns:
            CSV content as string
        """
        wb = load_workbook(io.BytesIO(excel_content), read_only=True, data_only=True)
        ws = wb.active

        output = io.StringIO()
        writer = csv.writer(output)

        for row in ws.iter_rows(values_only=True):
            # Skip empty rows
            if not any(cell for cell in row):
                continue
            writer.writerow(row)

        output.seek(0)
        return output.getvalue()

    def _identify_columns(self, headers: List[str]) -> Dict[str, str]:
        """
        Identify column names with flexible matching

        Args:
            headers: List of column headers from file

        Returns:
            Dictionary mapping standard names to actual column names
        """
        column_map = {}

        # Normalize headers for matching
        normalized_headers = {h.lower().strip(): h for h in headers}

        # Order Reference patterns
        order_ref_patterns = ['order reference', 'order ref', 'order id', '*order id', 'orderid']
        for pattern in order_ref_patterns:
            if pattern in normalized_headers:
                column_map['order_reference'] = normalized_headers[pattern]
                break

        # Reward ID patterns
        reward_id_patterns = ['reward id', 'reward_id', 'rewardid', 'id']
        for pattern in reward_id_patterns:
            if pattern in normalized_headers:
                column_map['reward_id'] = normalized_headers[pattern]
                break

        # Tracking ID patterns (required)
        tracking_id_patterns = ['tracking id', 'tracking_id', 'trackingid', 'awb', 'awb number']
        for pattern in tracking_id_patterns:
            if pattern in normalized_headers:
                column_map['tracking_id'] = normalized_headers[pattern]
                break

        # Tracking URL patterns (optional)
        tracking_url_patterns = ['tracking url', 'tracking_url', 'trackingurl', 'url']
        for pattern in tracking_url_patterns:
            if pattern in normalized_headers:
                column_map['tracking_url'] = normalized_headers[pattern]
                break

        # Courier Name patterns (optional)
        courier_patterns = ['courier name', 'courier_name', 'courier', 'carrier']
        for pattern in courier_patterns:
            if pattern in normalized_headers:
                column_map['courier_name'] = normalized_headers[pattern]
                break

        return column_map

    def _process_tracking_row(
        self,
        row: Dict[str, str],
        row_num: int,
        column_map: Dict[str, str],
        event_id: int,
        admin_id: int
    ) -> Dict:
        """
        Process a single row of tracking data

        Args:
            row: Row data as dictionary
            row_num: Row number (for error reporting)
            column_map: Column name mapping
            event_id: Event ID to filter
            admin_id: Admin user ID

        Returns:
            {
                success: bool,
                not_found: bool (optional),
                error: str (if failed),
                reward_id: str (if successful)
            }
        """
        # Extract values
        order_ref = row.get(column_map.get('order_reference', ''), '').strip()
        reward_id = row.get(column_map.get('reward_id', ''), '').strip()
        tracking_id = row.get(column_map['tracking_id'], '').strip()
        tracking_url = row.get(column_map.get('tracking_url', ''), '').strip()
        courier_name = row.get(column_map.get('courier_name', ''), '').strip()

        # Validate required data
        if not tracking_id:
            return {
                'success': False,
                'error': f"Row {row_num}: Tracking ID is empty"
            }

        if not order_ref and not reward_id:
            return {
                'success': False,
                'error': f"Row {row_num}: Both Order Reference and Reward ID are empty"
            }

        # Find reward
        query = self.db.query(UserReward).filter(UserReward.event_id == event_id)

        if order_ref:
            # Try finding by manual_order_reference first
            reward = query.filter(UserReward.manual_order_reference == order_ref).first()

            # If not found, try matching against generated reference pattern
            if not reward:
                # Generated reference format: RNR-EVT-{event_id}-USR-{user_id}-RWD-{reward_id_first_8}
                # Try to extract reward ID from order reference
                if order_ref.startswith('RNR-EVT-') and '-RWD-' in order_ref:
                    reward_id_part = order_ref.split('-RWD-')[-1][:8].upper()
                    # Find by partial UUID match
                    rewards = query.all()
                    for r in rewards:
                        if str(r.id)[:8].upper() == reward_id_part:
                            reward = r
                            break
        elif reward_id:
            reward = query.filter(UserReward.id == reward_id).first()

        if not reward:
            return {
                'success': False,
                'not_found': True,
                'error': f"Row {row_num}: Reward not found for order reference '{order_ref or reward_id}'"
            }

        # Validate reward status (should be READY_TO_SHIP)
        if reward.status not in [RewardStatus.READY_TO_SHIP, RewardStatus.TRACKING_ORDER]:
            return {
                'success': False,
                'error': (
                    f"Row {row_num}: Reward {reward.id} has invalid status '{reward.status.value}'. "
                    f"Expected 'ready_to_ship' or 'tracking_order'"
                )
            }

        # Update tracking fields
        reward.manual_tracking_id = tracking_id
        reward.manual_tracking_url = tracking_url if tracking_url else None
        reward.manual_courier_name = courier_name if courier_name else None
        reward.status = RewardStatus.TRACKING_ORDER
        reward.tracking_visible_to_user = True
        reward.tracking_imported_at = func.now()
        reward.tracking_imported_by_admin_id = admin_id

        # Store order reference if not already set
        if not reward.manual_order_reference:
            if order_ref:
                reward.manual_order_reference = order_ref
            else:
                # Generate and store reference
                reward.manual_order_reference = (
                    f"RNR-EVT-{reward.event_id}-USR-{reward.user_id}-RWD-{str(reward.id)[:8].upper()}"
                )

        logger.info(
            f"   ✓ Row {row_num}: Updated reward {reward.id} "
            f"with tracking ID {tracking_id}"
        )

        return {
            'success': True,
            'reward_id': str(reward.id)
        }
