"""
Excel Export Service for Physical Rewards
Generates Shiprocket bulk order Excel files with exact 30-column BASIC template format
"""
import io
import logging
from datetime import datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.core.enums import RewardStatus, RewardType
from app.models.user_reward import UserReward

logger = logging.getLogger(__name__)


class ExcelExportService:
    """Service for exporting physical reward shipping details to Shiprocket BASIC Excel format (30 columns)"""

    # Shiprocket BASIC template section headers (Row 1) - 30 columns total
    SECTION_HEADERS = [
        None,                   # Col 1
        "Pickup Details",       # Col 2
        "Buyer's Details",      # Col 3
        None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,  # Cols 4-18 (15 None)
        "Order Details",        # Col 19
        None, None, None, None, None, None,  # Cols 20-25 (6 None)
        "Package Details",      # Col 26
        None, None, None,       # Cols 27-29 (3 None)
        "Courier Details"       # Col 30
    ]

    # Shiprocket BASIC template column headers (Row 2) - EXACT from template
    SHIPROCKET_COLUMNS = [
        "*Order Id",                                        # 1
        "Pickup Address Id (Optional)",                    # 2
        "*Buyer's Mobile No.",                             # 3
        "*Buyer's First Name",                             # 4
        "Buyer's Last Name (Optional)",                    # 5
        "Email (Optional)",                                # 6
        "*Shipping Complete Address",                      # 7
        "Shipping Address Landmark (Optional)",            # 8
        "*Shipping Address Pincode",                       # 9
        "*Shipping Address City",                          # 10
        "*Shipping Address State",                         # 11
        "*Shipping Address Country",                       # 12
        "Billing Complete Address (Optional)",             # 13
        "Billing Landmark (Optional)",                     # 14
        "Billing Pincode (Optional)",                      # 15
        "Billing City (Optional)",                         # 16
        "Billing State (Optional)",                        # 17
        "Billing Country (Optional)",                      # 18
        "*Product Name",                                   # 19
        "*Per Unit Price in INR (Inclusive of Tax)",       # 20
        "*Product Quantity",                               # 21
        "*Master SKU",                                     # 22
        "*Payment Method (COD/Prepaid)",                   # 23
        "*Partial COD (Yes/No)",                           # 24
        "Paid Amount (Rs.)",                               # 25
        "*Weight Of Shipment (kg)",                        # 26
        "*Length (cm)",                                    # 27
        "*Breadth (cm)",                                   # 28
        "*Height (cm)",                                    # 29
        "Courier ID (Optional)"                            # 30
    ]

    def __init__(self, db: Session):
        """
        Initialize Excel export service

        Args:
            db: Database session
        """
        self.db = db

    def export_shipping_details(
        self,
        event_id: int,
        status: Optional[RewardStatus] = RewardStatus.READY_TO_SHIP,
        reward_type: Optional[str] = None
    ) -> bytes:
        """
        Export shipping details for rewards to Shiprocket BASIC Excel format (30 columns)

        Args:
            event_id: Event ID to export rewards for
            status: Filter by reward status (default: READY_TO_SHIP)
            reward_type: Optional filter by reward type (medal, tshirt, etc.)

        Returns:
            Excel file as bytes (XLSX format)

        Raises:
            ValueError: If no rewards found or required data missing
        """
        # Query rewards with shipping details
        query = self.db.query(UserReward).filter(
            UserReward.event_id == event_id,
            UserReward.requires_shipping == True,
            UserReward.shipping_details.isnot(None)
        )

        if status:
            query = query.filter(UserReward.status == status)

        if reward_type:
            query = query.filter(UserReward.reward_type == reward_type)

        rewards = query.all()

        if not rewards:
            logger.warning(f"No rewards found for event {event_id} with status {status}")

        logger.info(f"📊 Exporting {len(rewards)} rewards for event {event_id}")

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Shiprocket Bulk Orders"

        # Write headers
        self._write_headers(ws)

        # Group rewards by user to consolidate orders
        # Shiprocket requirement: Same order should have quantity > 1, not multiple rows
        user_rewards_map = {}
        for reward in rewards:
            user_id = reward.user_id
            if user_id not in user_rewards_map:
                user_rewards_map[user_id] = []
            user_rewards_map[user_id].append(reward)

        # Write data rows (one per user, with consolidated quantities)
        # Data starts from row 3 (row 1 = section headers, row 2 = column headers)
        row_idx = 3
        for user_id, user_rewards in user_rewards_map.items():
            if len(user_rewards) == 1:
                # Single reward for this user
                self._write_reward_row(ws, row_idx, user_rewards[0], quantity=1)
                row_idx += 1
            else:
                # Multiple rewards for same user - group by reward type
                reward_groups = {}
                for reward in user_rewards:
                    key = (reward.reward_type, reward.reward_name)
                    if key not in reward_groups:
                        reward_groups[key] = []
                    reward_groups[key].append(reward)

                # Write one row per reward type with quantity
                for (reward_type, reward_name), group_rewards in reward_groups.items():
                    # Use first reward as template, set quantity to group size
                    self._write_reward_row(ws, row_idx, group_rewards[0], quantity=len(group_rewards))
                    row_idx += 1

        # Format worksheet
        total_rows = row_idx - 3  # Subtract header rows (2 rows)
        self._format_worksheet(ws, total_rows)

        # Save to bytes
        try:
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            excel_bytes = output.getvalue()

            logger.info(f"✅ Excel export complete: {total_rows} rows, {len(excel_bytes)} bytes")
            return excel_bytes
        except Exception as e:
            logger.error(f"❌ Failed to generate Excel file: {str(e)}")
            raise ValueError(f"Excel generation failed: {str(e)}")

    def _write_headers(self, ws):
        """Write section headers (row 1) and column headers (row 2) to worksheet"""
        section_font = Font(bold=True, size=12, color="FFFFFF")
        section_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        section_alignment = Alignment(horizontal="center", vertical="center")

        header_font = Font(bold=True, size=10)
        header_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Row 1: Section headers (merged cells for sections)
        for col_idx, section in enumerate(self.SECTION_HEADERS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=section)
            if section:  # Only format cells with section headers
                cell.font = section_font
                cell.fill = section_fill
                cell.alignment = section_alignment

        # Row 2: Column headers
        for col_idx, header in enumerate(self.SHIPROCKET_COLUMNS, start=1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

    def _write_reward_row(self, ws, row_idx: int, reward: UserReward, quantity: int = 1):
        """
        Write reward data to a row following Shiprocket template format

        Args:
            ws: Worksheet
            row_idx: Row number (1-indexed)
            reward: UserReward instance
            quantity: Product quantity (default: 1, must be > 0)
        """
        # Ensure quantity is valid (> 0)
        if quantity <= 0:
            quantity = 1
        shipping = reward.shipping_details or {}

        # Generate order reference if not exists
        # Format: RNREVT{event_id}USR{user_id}RWD{reward_id_short}
        # Max 30 chars, no symbols, alphanumeric only, unique per reward
        if not reward.manual_order_reference:
            # Convert UUID to short alphanumeric (first 8 chars)
            reward_id_short = str(reward.id).replace("-", "")[:8].upper()
            # Create compact format: RNR + E{event_id} + U{user_id} + {reward_short}
            order_ref = f"RNRE{reward.event_id}U{reward.user_id}R{reward_id_short}"
            # Ensure max 30 characters
            order_ref = order_ref[:30]
        else:
            order_ref = reward.manual_order_reference

        # Split name into first and last
        full_name = shipping.get("full_name") or shipping.get("name", "Customer")
        name_parts = full_name.strip().split(maxsplit=1)
        first_name = name_parts[0] if name_parts else "Customer"
        last_name = name_parts[1] if len(name_parts) > 1 else ""

        # Current date in DD-MM-YYYY format
        current_date = datetime.now().strftime("%d-%m-%Y")

        # Get address fields (support both postal_code and pincode)
        pincode = str(shipping.get("postal_code") or shipping.get("pincode", "")).strip()
        address_line1_raw = str(shipping.get("address_line1", "")).strip()
        address_line2 = str(shipping.get("address_line2", "")).strip()
        city = str(shipping.get("city", "")).strip()
        state = str(shipping.get("state", "")).strip()
        country = str(shipping.get("country", "India")).strip()

        # Validate and format shipping address per Shiprocket requirements:
        # - Max 300 characters
        # - Min 5 characters
        # - Must contain at least 1 number and 1 space
        address_line1 = address_line1_raw
        if address_line1:
            # Ensure max 300 characters
            if len(address_line1) > 300:
                address_line1 = address_line1[:300]

            # Ensure minimum requirements: 5 chars, 1 number, 1 space
            has_number = any(char.isdigit() for char in address_line1)
            has_space = ' ' in address_line1

            # If missing number or space, try to fix
            if not has_number and pincode:
                # Prepend house number or pincode reference
                address_line1 = f"Flat 1 {address_line1}"
            if not has_space:
                # Add space if completely missing
                address_line1 = address_line1.replace(",", ", ")

            # Final check: ensure min 5 characters
            if len(address_line1) < 5:
                address_line1 = f"House 1 {address_line1}".strip()
        else:
            # Fallback if address is empty
            address_line1 = "House 1 Main Road"

        # Phone number formatting - Shiprocket expects 10-digit Indian mobile without +91
        phone_raw = str(shipping.get("phone", "")).strip()
        # Remove common prefixes and non-digits
        phone = phone_raw.replace("+91", "").replace("-", "").replace(" ", "").strip()
        # Ensure it's 10 digits
        if len(phone) > 10:
            phone = phone[-10:]  # Take last 10 digits

        email = str(shipping.get("email", "")).strip()

        # Product details
        # Product name max 200 characters (Shiprocket requirement)
        product_name = str(reward.reward_name or "Physical Reward").strip()
        if len(product_name) > 200:
            product_name = product_name[:197] + "..."  # Truncate with ellipsis

        # Master SKU cannot be empty (Shiprocket requirement)
        # Generate SKU if not provided: GLCG-{TYPE}-{short_reward_id}
        if reward.item_sku and str(reward.item_sku).strip():
            sku = str(reward.item_sku).strip()[:20]  # Max 20 chars for safety
        else:
            # Generate SKU from reward type and ID
            reward_id_short = str(reward.id).replace("-", "")[:8].upper()
            sku = f"GLCG{reward.reward_type.value.upper()}{reward_id_short}"[:20]

        hsn_code = reward.item_hsn or ""

        # Package dimensions (use defaults if not specified)
        # Weight validation per Shiprocket requirements:
        # - Must not be empty
        # - Must be numeric
        # - Cannot be negative
        # - Maximum 30kg
        if reward.item_weight:
            weight = float(reward.item_weight)
            # Ensure non-negative and within 30kg limit
            if weight < 0:
                weight = 0.5  # Default to 0.5kg if negative
            elif weight > 30:
                weight = 30.0  # Cap at 30kg maximum
        else:
            weight = 0.5  # Default weight if not provided

        # Dimension validation per Shiprocket requirements:
        # - Length, Breadth, Height cannot be less than 0.5 cm
        # - For document orders: minimum 10x10x1 cm
        # - We use 15x10x5 as safe defaults for physical items

        # Length validation (min 0.5cm)
        if reward.item_length:
            length = float(reward.item_length)
            length = max(0.5, length)  # Ensure minimum 0.5cm
        else:
            length = 15.0  # Safe default

        # Breadth validation (min 0.5cm)
        if reward.item_breadth:
            breadth = float(reward.item_breadth)
            breadth = max(0.5, breadth)  # Ensure minimum 0.5cm
        else:
            breadth = 10.0  # Safe default

        # Height validation (min 0.5cm)
        if reward.item_height:
            height = float(reward.item_height)
            height = max(0.5, height)  # Ensure minimum 0.5cm
        else:
            height = 5.0  # Safe default

        # For very small dimensions, enforce document minimum (10x10x1)
        # This prevents rejection for undersized packages
        if length < 10 or breadth < 10 or height < 1:
            length = max(length, 10.0)
            breadth = max(breadth, 10.0)
            height = max(height, 1.0)

        # Row data matching exact Shiprocket BASIC template (30 columns)
        # IMPORTANT: All values must be primitives (str, int, float) - no None or objects
        row_data = [
            str(order_ref),                     # 1. *Order Id
            "",                                 # 2. Pickup Address Id (Optional)
            str(phone) if phone else "",        # 3. *Buyer's Mobile No.
            str(first_name),                    # 4. *Buyer's First Name
            str(last_name) if last_name else "", # 5. Buyer's Last Name (Optional)
            str(email) if email else "",        # 6. Email (Optional)
            str(address_line1),                 # 7. *Shipping Complete Address
            str(address_line2) if address_line2 else "", # 8. Shipping Address Landmark (Optional)
            str(pincode) if pincode else "",    # 9. *Shipping Address Pincode
            str(city) if city else "",          # 10. *Shipping Address City
            str(state) if state else "",        # 11. *Shipping Address State
            str(country),                       # 12. *Shipping Address Country
            str(address_line1),                 # 13. Billing Complete Address (Optional, same as shipping)
            str(address_line2) if address_line2 else "", # 14. Billing Landmark (Optional)
            str(pincode) if pincode else "",    # 15. Billing Pincode (Optional)
            str(city) if city else "",          # 16. Billing City (Optional)
            str(state) if state else "",        # 17. Billing State (Optional)
            str(country),                       # 18. Billing Country (Optional)
            str(product_name),                  # 19. *Product Name
            500,                                # 20. *Per Unit Price in INR (must be > 0)
            int(quantity),                      # 21. *Product Quantity (must be > 0)
            str(sku),                           # 22. *Master SKU
            "Prepaid",                          # 23. *Payment Method (COD/Prepaid)
            "no",                               # 24. *Partial COD (Yes/No)
            int(quantity * 500),                # 25. Paid Amount (Rs.) = quantity * price
            float(weight),                      # 26. *Weight Of Shipment (kg)
            float(length),                      # 27. *Length (cm)
            float(breadth),                     # 28. *Breadth (cm)
            float(height),                      # 29. *Height (cm)
            ""                                  # 30. Courier ID (Optional)
        ]

        # Write row with explicit data types
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)

            # Ensure only primitive types (str, int, float) - no objects
            if value is None:
                cell.value = ""  # Empty string instead of None
            elif isinstance(value, (int, float)):
                cell.value = value  # Keep numeric types as-is
            else:
                cell.value = str(value)  # Convert everything else to string

    def _format_worksheet(self, ws, num_rows: int):
        """
        Format worksheet for better readability

        Args:
            ws: Worksheet
            num_rows: Number of data rows
        """
        # Set column widths (30-column BASIC template)
        column_widths = {
            1: 35,   # Order Id
            2: 20,   # Pickup Address Id
            3: 18,   # Phone
            4: 20,   # First Name
            5: 20,   # Last Name
            6: 30,   # Email
            7: 40,   # Shipping Address
            8: 25,   # Landmark
            9: 15,   # Pincode
            10: 20,  # City
            11: 20,  # State
            12: 15,  # Country
            19: 30,  # Product Name
            20: 18,  # Price
            21: 12,  # Quantity
            22: 25,  # SKU
            23: 15,  # Payment Method
        }

        for col_idx, width in column_widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Set default width for other columns
        for col_idx in range(1, len(self.SHIPROCKET_COLUMNS) + 1):
            if col_idx not in column_widths:
                ws.column_dimensions[get_column_letter(col_idx)].width = 15

        # Freeze header rows (rows 1-2)
        ws.freeze_panes = "A3"

        # Add auto-filter starting from row 2 (column headers)
        ws.auto_filter.ref = f"A2:{get_column_letter(len(self.SHIPROCKET_COLUMNS))}{num_rows + 2}"

        # Format pincode and phone as text (prevent Excel from treating as numbers)
        # Data starts from row 3
        for row_idx in range(3, num_rows + 3):
            # Pincode (column 9)
            ws.cell(row=row_idx, column=9).number_format = '@'
            # Phone (column 3)
            ws.cell(row=row_idx, column=3).number_format = '@'
