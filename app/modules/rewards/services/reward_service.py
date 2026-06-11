"""
Reward Service

Business logic for physical reward fulfillment via Shiprocket.
"""

import io
import logging
from datetime import datetime
from typing import Optional

import openpyxl
from fastapi import HTTPException, UploadFile, status
from openpyxl.styles import Font, PatternFill
from sqlalchemy import and_, or_
from sqlalchemy.orm import Session, joinedload

from app.core.enums import RewardStatus
from app.core.exceptions import (
    AlreadyExistsException,
    NotFoundException,
)
from app.models.user import User
from app.models.user_reward import RewardType, UserReward
from app.modules.events.domain.event import Event
from app.modules.registrations.domain.registration import Registration
from app.modules.rewards.domain.value_objects import (
    ShippingAddress,
)
from app.modules.rewards.schemas.reward import BulkShipmentUpdateResponse
from app.services.base import BaseService

logger = logging.getLogger(__name__)


class RewardService(BaseService):
    """Service for physical reward operations"""

    def __init__(self, db: Session):
        super().__init__(db)

    def create_reward_order(
        self, registration_id: int, reward_name: str, shipping_address: ShippingAddress
    ) -> UserReward:
        """
        Create reward order.

        Business Rules:
        1. Registration must exist
        2. One reward per registration
        3. Shipping address must be valid

        Args:
            registration_id: Registration ID
            reward_name: Name of reward
            shipping_address: Shipping address

        Returns:
            Created UserReward

        Raises:
            NotFoundException: If registration not found
            AlreadyExistsException: If reward already exists
        """
        # Get registration
        registration = (
            self.db.query(Registration).filter(Registration.id == registration_id).first()
        )

        if not registration:
            raise NotFoundException("Registration", str(registration_id))

        # Check if reward already exists
        existing = (
            self.db.query(UserReward)
            .filter(
                and_(
                    UserReward.registration_id == registration_id,
                    UserReward.reward_type == RewardType.MEDAL,
                )
            )
            .first()
        )

        if existing:
            raise AlreadyExistsException("Reward", "registration_id", str(registration_id))

        # Create reward record
        reward = UserReward(
            user_id=registration.user_id,
            registration_id=registration_id,
            event_id=registration.event_id,
            reward_id=f"medal-{registration_id}",
            reward_type=RewardType.MEDAL,
            reward_name=reward_name,
            status=RewardStatus.LOCKED.value,
        )

        self.db.add(reward)
        self.db.commit()
        self.db.refresh(reward)

        # TODO: Create Shiprocket order via external service
        # shiprocket_service.create_order(reward, shipping_address)

        return reward

    def update_shipment_status(
        self,
        reward_id: int,
        status: str,
        tracking_number: str | None = None,
        shiprocket_order_id: str | None = None,
    ) -> UserReward:
        """Update shipment status. `status` should be a RewardStatus value string."""
        reward = self.db.query(UserReward).filter(UserReward.id == reward_id).first()
        if not reward:
            raise NotFoundException("Reward", str(reward_id))

        reward.status = status

        if tracking_number:
            reward.tracking_number = tracking_number

        if shiprocket_order_id:
            reward.shiprocket_order_id = shiprocket_order_id

        self.db.commit()
        self.db.refresh(reward)

        return reward

    def get_user_rewards(self, user_id: int) -> list[UserReward]:
        """Get all physical rewards for user"""
        return (
            self.db.query(UserReward)
            .filter(and_(UserReward.user_id == user_id, UserReward.reward_type == RewardType.MEDAL))
            .order_by(UserReward.created_at.desc())
            .all()
        )

    def get_pending_rewards(self) -> list[UserReward]:
        """Get all pending reward orders"""
        return (
            self.db.query(UserReward)
            .filter(
                and_(
                    UserReward.reward_type == RewardType.MEDAL,
                    UserReward.status == RewardStatus.LOCKED.value,
                )
            )
            .all()
        )

    def admin_unlock_reward(
        self, event_id: int, user_id: int, registration_id: int, admin_id: int | None = None
    ) -> UserReward:
        """
        Admin verifies shipping and unlocks reward (idempotent operation).

        IDEMPOTENT: Can be called multiple times without errors.
        - If reward doesn't exist: Creates new UserReward with is_verified=True
        - If reward exists: Updates is_verified=True and shipping_details from registration

        Business Rules:
        1. Registration must exist and match event_id/user_id
        2. Updates existing reward OR creates new one (idempotent)
        3. If shipping details exist in registration:
           - Reward status: 'READY_TO_SHIP' (ready to ship)
           - shipping_details populated from registration
           - Attempts Shiprocket order creation (non-blocking)
        4. If shipping details missing:
           - Reward status: 'PENDING_DETAILS' (needs address)
        5. Preserves SHIPPED/DELIVERED status (doesn't downgrade)

        Args:
            event_id: Event ID
            user_id: User ID
            registration_id: Registration ID
            admin_id: Admin user ID (optional, for tracking who verified)

        Returns:
            UserReward (created or updated)

        Raises:
            NotFoundException: If registration not found
        """
        # Get registration and validate it belongs to user/event
        registration = (
            self.db.query(Registration)
            .filter(
                and_(
                    Registration.id == registration_id,
                    Registration.user_id == user_id,
                    Registration.event_id == event_id,
                )
            )
            .first()
        )

        if not registration:
            raise NotFoundException("Registration", f"id={registration_id}, user_id={user_id}, event_id={event_id}")

        # Check if reward already exists
        existing = (
            self.db.query(UserReward)
            .filter(
                and_(
                    UserReward.registration_id == registration_id,
                    UserReward.reward_type == RewardType.MEDAL,
                )
            )
            .first()
        )

        if existing:
            # UPDATE EXISTING REWARD - RE-VERIFICATION FLOW
            logger.info(f"Reward already exists for registration {registration_id}, updating verification status")

            # Update verification status
            existing.is_verified = True
            existing.verified_by_admin_id = admin_id
            existing.verified_at = datetime.utcnow()
            existing.is_unlocked = True  # Ensure reward is unlocked
            existing.unlocked_by_admin_id = admin_id

            # Check if registration has complete shipping details (reuse logic below)
            has_shipping = bool(
                registration.shipping_address_line1
                and registration.shipping_city
                and registration.shipping_state
                and registration.shipping_postal_code
                and registration.shipping_phone
            )

            # Update shipping details from registration
            if has_shipping:
                shipping_details = {
                    "full_name": registration.participant_name or "Unknown",
                    "phone": registration.shipping_phone,
                    "address_line1": registration.shipping_address_line1,
                    "address_line2": registration.shipping_address_line2 or "",
                    "city": registration.shipping_city,
                    "state": registration.shipping_state,
                    "postal_code": registration.shipping_postal_code,
                    "country": registration.shipping_country or "India",
                    "email": registration.shipping_email or registration.user.email if registration.user else "",
                }
                existing.shipping_details = shipping_details

                # Only update status if not already tracking/delivered
                if existing.status not in [RewardStatus.TRACKING_ORDER.value, RewardStatus.DELIVERED.value]:
                    existing.status = RewardStatus.READY_TO_SHIP.value
                logger.info(f"Updated shipping details for existing reward {existing.id}")
            else:
                # Shipping incomplete - set to LOCKED
                if existing.status not in [RewardStatus.TRACKING_ORDER.value, RewardStatus.DELIVERED.value]:
                    existing.status = RewardStatus.LOCKED.value
                logger.info(f"Shipping incomplete, status set to LOCKED for reward {existing.id}")

            # Commit and refresh
            self.db.commit()
            self.db.refresh(existing)

            logger.info(f"Re-verified existing reward {existing.id} for registration {registration_id}")

            # Try to create/update Shiprocket order if shipping is complete
            if has_shipping:
                from app.modules.shipping.integrations.shiprocket.fulfillment_service import RewardFulfillmentService

                fulfillment_service = RewardFulfillmentService(self.db)

                try:
                    logger.info(f"Creating/updating Shiprocket pre-order for reward {existing.id}")
                    fulfillment_service.create_shiprocket_order(str(existing.id))
                    logger.info(f"✅ Shiprocket pre-order created/updated successfully for reward {existing.id}")
                except Exception as e:
                    error_message = f"Shiprocket order creation failed: {str(e)}"
                    logger.warning(f"⚠️ {error_message} for reward {existing.id}")
                    existing.fulfillment_error = error_message
                    self.db.commit()
                    self.db.refresh(existing)

            return existing

        # Check if registration has complete shipping details
        has_shipping = bool(
            registration.shipping_address_line1
            and registration.shipping_city
            and registration.shipping_state
            and registration.shipping_postal_code
            and registration.shipping_phone
        )

        # If shipping details exist, populate them immediately and set status to pending_shipment
        shipping_details = None
        reward_status = RewardStatus.LOCKED.value  # Default: waiting for user

        if has_shipping:
            # Populate shipping_details JSONB immediately (admin verified)
            # Use "full_name" and "postal_code" to match UserReward model documentation
            shipping_details = {
                "full_name": registration.participant_name or "Unknown",
                "phone": registration.shipping_phone,
                "address_line1": registration.shipping_address_line1,
                "address_line2": registration.shipping_address_line2 or "",
                "city": registration.shipping_city,
                "state": registration.shipping_state,
                "postal_code": registration.shipping_postal_code,
                "country": registration.shipping_country or "India",
                "email": registration.shipping_email or registration.user.email if registration.user else "",
            }
            reward_status = RewardStatus.READY_TO_SHIP.value  # Ready to ship

        # Create reward record in unlocked state
        reward = UserReward(
            user_id=user_id,
            registration_id=registration_id,
            event_id=event_id,
            reward_id=f"medal-{registration_id}",
            reward_type=RewardType.MEDAL,
            reward_name=f"{registration.current_tier.tier_name} Medal" if registration.current_tier else "Event Medal",
            status=reward_status,
            shipping_details=shipping_details,
            is_unlocked=True,  # Admin unlock - set to True
            unlocked_by_admin_id=admin_id,
            is_verified=True,  # Admin verified shipping
            verified_by_admin_id=admin_id,
            verified_at=datetime.utcnow(),
        )

        self.db.add(reward)
        self.db.commit()
        self.db.refresh(reward)

        logger.info(f"Admin unlocked reward for user_id={user_id}, event_id={event_id}, registration_id={registration_id}")

        # If shipping details are complete, automatically create Shiprocket pre-order
        if has_shipping:
            from app.modules.shipping.integrations.shiprocket.fulfillment_service import RewardFulfillmentService

            fulfillment_service = RewardFulfillmentService(self.db)

            try:
                logger.info(f"Creating Shiprocket pre-order for reward {reward.id}")
                # Create Shiprocket order (non-blocking - errors won't fail unlock)
                fulfillment_service.create_shiprocket_order(str(reward.id))
                logger.info(f"✅ Shiprocket pre-order created successfully for reward {reward.id}")
            except Exception as e:
                # Non-blocking error handling - reward is still unlocked, admin can retry later
                error_message = f"Shiprocket order creation failed: {str(e)}"
                logger.warning(f"⚠️ {error_message} for reward {reward.id}")

                # Store error in fulfillment_error field so admin can see it
                reward.fulfillment_error = error_message
                self.db.commit()
                self.db.refresh(reward)

        return reward

    def toggle_tracking_visibility(self, reward_id: str, visible: bool) -> UserReward:
        """
        Toggle tracking visibility for user.

        Args:
            reward_id: UUID of reward
            visible: True to show tracking, False to hide

        Returns:
            Updated UserReward

        Raises:
            NotFoundException: If reward not found
            ValueError: If trying to show tracking when no tracking number exists
        """
        from uuid import UUID

        # Convert string to UUID
        try:
            reward_uuid = UUID(reward_id)
        except ValueError:
            raise ValueError(f"Invalid reward ID format: {reward_id}")

        reward = self.db.query(UserReward).filter(UserReward.id == reward_uuid).first()

        if not reward:
            raise NotFoundException("Reward", "id", str(reward_id))

        # Validate tracking exists before allowing unlock (check both legacy and manual tracking)
        if visible and not (reward.tracking_number or reward.manual_tracking_url):
            raise ValueError("Cannot show tracking - no tracking number or URL available yet")

        reward.tracking_visible_to_user = visible
        self.db.commit()
        self.db.refresh(reward)

        logger.info(
            f"Tracking visibility {'enabled' if visible else 'disabled'} for reward {reward_id}"
        )

        return reward

    def get_all_rewards_with_details(
        self,
        status_filter: Optional[str] = None,
        search: Optional[str] = None,
    ) -> list[dict]:
        """
        Get all rewards with full details for admin dashboard.

        Args:
            status_filter: Filter by reward status
            search: Search by user name, email, or tracking number

        Returns:
            List of reward dicts with all details
        """
        # Build query with joins
        query = (
            self.db.query(UserReward)
            .join(User, UserReward.user_id == User.id)
            .join(Registration, UserReward.registration_id == Registration.id)
            .join(Event, UserReward.event_id == Event.id)
            .options(
                joinedload(UserReward.user),
                joinedload(UserReward.registration),
                joinedload(UserReward.event),
            )
        )

        # Apply status filter
        if status_filter:
            query = query.filter(UserReward.status == status_filter)

        # Apply search filter
        if search:
            search_term = f"%{search}%"
            query = query.filter(
                or_(
                    User.first_name.ilike(search_term),
                    User.last_name.ilike(search_term),
                    User.email.ilike(search_term),
                    UserReward.tracking_number.ilike(search_term),
                )
            )

        rewards = query.order_by(UserReward.created_at.desc()).all()

        # Transform to dict with all details
        results = []
        for reward in rewards:
            results.append({
                # Reward fields
                "id": reward.id,
                "reward_id": reward.reward_id,
                "reward_name": reward.reward_name,
                "reward_type": reward.reward_type.value if hasattr(reward.reward_type, 'value') else reward.reward_type,
                "status": reward.status,
                # User details
                "user_id": reward.user.id,
                "user_name": reward.user.full_name,
                "user_email": reward.user.email,
                "user_phone": reward.user.phone if hasattr(reward.user, 'phone') else None,
                # Event details
                "event_id": reward.event.id,
                "event_name": reward.event.name,
                # Registration details
                "registration_id": reward.registration.id,
                "registration_number": reward.registration.registration_number,
                "tier_name": reward.registration.current_tier.tier_name if reward.registration.current_tier else None,
                # Shipping details
                "shipping_address": reward.shipping_details if hasattr(reward, 'shipping_details') else None,
                "tracking_number": reward.tracking_number,
                "courier_partner": reward.courier_partner,
                "shiprocket_order_id": reward.shiprocket_order_id,
                "shiprocket_shipment_id": reward.shiprocket_shipment_id,
                # Timestamps
                "created_at": reward.created_at,
                "shipped_at": reward.shipped_at,
                "delivered_at": reward.delivered_at,
                "estimated_delivery": reward.estimated_delivery_date if hasattr(reward, 'estimated_delivery_date') else None,
                # Progress (from registration)
                "total_distance_km": reward.registration.total_distance_km if hasattr(reward.registration, 'total_distance_km') else None,
                "goal_distance_km": reward.registration.current_tier.goal_distance_km if reward.registration.current_tier and hasattr(reward.registration.current_tier, 'goal_distance_km') else None,
                "progress_percentage": (
                    int((reward.registration.total_distance_km / reward.registration.current_tier.goal_distance_km) * 100)
                    if reward.registration.current_tier
                    and hasattr(reward.registration, 'total_distance_km')
                    and hasattr(reward.registration.current_tier, 'goal_distance_km')
                    and reward.registration.current_tier.goal_distance_km > 0
                    else 0
                ),
                "proof_image_url": reward.registration.proof_image_url if hasattr(reward.registration, 'proof_image_url') else None,
            })

        return results

    async def get_shipping_preview(self, reward_id: str) -> dict:
        """
        Get shipping preview with all details before creating order.

        Shows admin:
        - Reward details (name, type)
        - Package dimensions and weight
        - Full shipping address and phone
        - Pickup location details
        - Available couriers with estimated costs
        - Estimated delivery time
        - Serviceability check

        Args:
            reward_id: Reward ID (UUID string)

        Returns:
            Complete shipping preview data

        Raises:
            NotFoundException: If reward not found
            ValueError: If reward not ready for preview
        """
        # Get reward with shipping details
        from uuid import UUID
        try:
            reward_uuid = UUID(reward_id)
        except ValueError:
            raise ValueError(f"Invalid reward ID format: {reward_id}")

        reward = self.db.query(UserReward).filter(UserReward.id == reward_uuid).first()
        if not reward:
            raise NotFoundException("Reward", str(reward_id))

        # Validate shipping address exists
        if not reward.shipping_details:
            raise ValueError("Reward does not have shipping address. User must provide shipping details first.")

        # Get Shiprocket config for default dimensions
        from app.modules.shipping.domain.config import ShiprocketConfig
        config = self.db.query(ShiprocketConfig).filter(ShiprocketConfig.is_active).first()
        if not config:
            raise ValueError("Shiprocket configuration not found.")

        # Parse shipping address
        shipping_addr = reward.shipping_details
        delivery_pincode = shipping_addr.get("pincode", "")

        # Check serviceability and get courier rates
        from app.modules.shipping.integrations.shiprocket.client import ShiprocketService
        shiprocket = ShiprocketService(self.db)

        serviceability = await shiprocket.check_pincode_serviceability(
            delivery_pincode=delivery_pincode,
            weight=float(reward.item_weight or config.default_weight)
        )

        # Format available couriers with rates
        available_couriers = []
        if serviceability.get("success") and serviceability.get("available_couriers"):
            for courier in serviceability["available_couriers"][:5]:  # Top 5 couriers
                available_couriers.append({
                    "name": courier.get("courier_name", "Unknown"),
                    "rate": courier.get("rate", 0),
                    "etd": courier.get("etd", "N/A"),
                    "cod_available": courier.get("cod", 0) == 1,
                })

        # Format addresses
        shipping_address_line = shipping_addr.get("address_line1", "")
        if shipping_addr.get("address_line2"):
            shipping_address_line += f", {shipping_addr.get('address_line2')}"

        pickup_address_line = f"Gahlot House, Ground Floor, Gyan Sarover Colony, Tiraya, Rajasthan 324008"

        return {
            "reward_name": reward.reward_name,
            "reward_type": reward.reward_type.value,
            # Package details
            "length_cm": float(reward.item_length or config.default_length),
            "breadth_cm": float(reward.item_breadth or config.default_breadth),
            "height_cm": float(reward.item_height or config.default_height),
            "weight_kg": float(reward.item_weight or config.default_weight),
            # Shipping address
            "shipping_name": shipping_addr.get("name", ""),
            "shipping_address": shipping_address_line,
            "shipping_city": shipping_addr.get("city", ""),
            "shipping_state": shipping_addr.get("state", ""),
            "shipping_pincode": delivery_pincode,
            "shipping_phone": shipping_addr.get("phone", ""),
            # Pickup location
            "pickup_location": config.default_pickup_location,
            "pickup_address": pickup_address_line,
            # Serviceability
            "available_couriers": available_couriers if available_couriers else None,
            "estimated_delivery_days": available_couriers[0]["etd"] if available_couriers else "N/A",
            "is_serviceable": serviceability.get("is_serviceable", False),
        }

    async def ship_reward_with_shiprocket(
        self,
        reward_id: str,
        courier_id: Optional[int] = None,
        selection_strategy: Optional[str] = None
    ) -> dict:
        """
        Ship reward automatically using Shiprocket with optional courier override.

        Args:
            reward_id: Reward ID (UUID string)
            courier_id: Optional courier company ID for manual selection (overrides auto-selection)
            selection_strategy: Optional strategy override ('cheapest', 'fastest', 'balanced')

        Returns:
            Shiprocket shipment details including courier selection metadata

        Raises:
            NotFoundException: If reward not found
            ValueError: If reward not ready for shipping
        """
        # Get reward
        from uuid import UUID
        try:
            reward_uuid = UUID(reward_id)
        except ValueError:
            raise ValueError(f"Invalid reward ID format: {reward_id}")

        reward = self.db.query(UserReward).filter(UserReward.id == reward_uuid).first()
        if not reward:
            raise NotFoundException("Reward", str(reward_id))

        # Validate status
        if reward.status != RewardStatus.READY_TO_SHIP.value:
            raise ValueError(f"Reward must be in 'pending_shipment' status to ship. Current status: {reward.status}")

        # Validate shipping address exists
        if not reward.shipping_details:
            raise ValueError("Reward does not have shipping address. User must provide shipping details first.")

        # Initialize fulfillment service
        from app.modules.shipping.integrations.shiprocket.fulfillment_service import RewardFulfillmentService
        fulfillment = RewardFulfillmentService(self.db)

        # Create Shiprocket order
        logger.info(f"Creating Shiprocket order for reward_id={reward_id}")
        order_result = await fulfillment.create_shiprocket_order(str(reward.id))

        # Check if order creation succeeded
        if not order_result.get("success"):
            error_msg = order_result.get("error", "Unknown error creating Shiprocket order")
            logger.error(f"❌ Shiprocket order creation failed for reward_id={reward_id}: {error_msg}")
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Failed to create Shiprocket order: {error_msg}"
            )

        # Assign AWB and generate label with courier selection
        logger.info(
            f"Assigning AWB for reward_id={reward_id} "
            f"(courier_id={courier_id}, strategy={selection_strategy})"
        )
        awb_result = await fulfillment.assign_awb_and_generate_label(
            reward_id=str(reward.id),
            courier_id=courier_id,
            selection_strategy=selection_strategy
        )

        # Schedule pickup
        logger.info(f"Scheduling pickup for reward_id={reward_id}")
        pickup_result = await fulfillment.schedule_pickup(str(reward.id))

        # Refresh reward to get updated data
        self.db.refresh(reward)

        return {
            "success": True,
            "tracking_number": reward.tracking_number,
            "courier_partner": reward.courier_partner,
            "awb": reward.tracking_number,  # Same as tracking_number
            "label_url": awb_result.get("label_url", ""),
            "shiprocket_order_id": int(reward.shiprocket_order_id) if reward.shiprocket_order_id else 0,
            "shiprocket_shipment_id": int(reward.shiprocket_shipment_id) if reward.shiprocket_shipment_id else 0,
            "pickup_scheduled_date": pickup_result.get("pickup_scheduled_date"),
            # Add courier selection metadata
            "courier_id": awb_result.get("courier_id"),
            "selected_courier_rate": awb_result.get("selected_courier_rate"),
            "cost_savings": awb_result.get("cost_savings"),
            "selection_strategy_used": awb_result.get("selection_strategy_used"),
        }

    def ship_reward_manually(
        self,
        reward_id: str,
        tracking_number: str,
        courier_partner: str,
        shipped_at: Optional[datetime] = None,
    ) -> UserReward:
        """
        Mark reward as shipped with manual tracking details.

        Args:
            reward_id: Reward ID (UUID string)
            tracking_number: Tracking/AWB number
            courier_partner: Courier company name
            shipped_at: Optional ship timestamp (defaults to now)

        Returns:
            Updated UserReward

        Raises:
            NotFoundException: If reward not found
        """
        from uuid import UUID
        try:
            reward_uuid = UUID(reward_id)
        except ValueError:
            raise ValueError(f"Invalid reward ID format: {reward_id}")

        reward = self.db.query(UserReward).filter(UserReward.id == reward_uuid).first()
        if not reward:
            raise NotFoundException("Reward", str(reward_id))

        # Update reward with shipping details
        reward.status = RewardStatus.SHIPPED.value
        reward.tracking_number = tracking_number
        reward.courier_partner = courier_partner
        reward.shipped_at = shipped_at or datetime.utcnow()

        # Log in status history if it exists
        if hasattr(reward, 'status_history') and reward.status_history:
            history = reward.status_history if isinstance(reward.status_history, list) else []
            history.append({
                "status": "shipped",
                "timestamp": (shipped_at or datetime.utcnow()).isoformat(),
                "source": "manual",
                "note": f"Shipped manually by admin via {courier_partner}",
            })
            reward.status_history = history

        self.db.commit()
        self.db.refresh(reward)

        logger.info(f"Reward {reward_id} marked as shipped manually. Tracking: {tracking_number}, Courier: {courier_partner}")

        return reward

    def mark_reward_delivered(self, reward_id: str) -> UserReward:
        """
        Mark reward as delivered.

        Args:
            reward_id: Reward ID (UUID string)

        Returns:
            Updated UserReward

        Raises:
            NotFoundException: If reward not found
        """
        from uuid import UUID
        try:
            reward_uuid = UUID(reward_id)
        except ValueError:
            raise ValueError(f"Invalid reward ID format: {reward_id}")

        reward = self.db.query(UserReward).filter(UserReward.id == reward_uuid).first()
        if not reward:
            raise NotFoundException("Reward", str(reward_id))

        # Update status to delivered
        reward.status = RewardStatus.DELIVERED.value
        reward.delivered_at = datetime.utcnow()

        # Log in status history if it exists
        if hasattr(reward, 'status_history') and reward.status_history:
            history = reward.status_history if isinstance(reward.status_history, list) else []
            history.append({
                "status": "delivered",
                "timestamp": datetime.utcnow().isoformat(),
                "source": "manual",
                "note": "Marked as delivered manually by admin",
            })
            reward.status_history = history

        self.db.commit()
        self.db.refresh(reward)

        logger.info(f"Reward {reward_id} marked as delivered")

        return reward

    def export_pending_shipments_to_excel(self, event_id: Optional[int] = None) -> io.BytesIO:
        """
        Export all pending shipments to Excel for Shiprocket bulk upload.

        Args:
            event_id: Optional event ID filter

        Returns:
            BytesIO buffer containing Excel file
        """
        # Query pending shipments
        query = self.db.query(UserReward).filter(
            UserReward.status == RewardStatus.READY_TO_SHIP.value,
            UserReward.shipping_details.isnot(None),
        )

        if event_id:
            query = query.filter(UserReward.event_id == event_id)

        rewards = query.order_by(UserReward.created_at.asc()).all()

        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Pending Shipments"

        # Define headers
        headers = [
            "Internal ID",
            "Order Reference",
            "Full Name",
            "Address Line 1",
            "Address Line 2",
            "City",
            "State",
            "Pincode",
            "Phone",
            "Email",
            "Product Name",
            "Product SKU",
            "Weight (kg)",
            "Length (cm)",
            "Breadth (cm)",
            "Height (cm)",
            "AWB Code",
            "Tracking Number",
            "Courier Name",
        ]

        # Style header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font

        # Fill data rows
        for row_num, reward in enumerate(rewards, 2):
            shipping = reward.shipping_details or {}

            # Generate order reference
            order_ref = f"RNR-EVT-{reward.event_id}-USR-{reward.user_id}-RWD-{str(reward.id)[:8].upper()}"

            # Support both field name formats
            full_name = shipping.get("full_name") or shipping.get("name", "")
            postal_code = shipping.get("postal_code") or shipping.get("pincode", "")

            row_data = [
                str(reward.id),  # Internal ID
                order_ref,  # Order Reference
                full_name,  # Full Name
                shipping.get("address_line1", ""),  # Address Line 1
                shipping.get("address_line2", ""),  # Address Line 2
                shipping.get("city", ""),  # City
                shipping.get("state", ""),  # State
                postal_code,  # Pincode
                shipping.get("phone", ""),  # Phone
                shipping.get("email", ""),  # Email
                reward.reward_name or "Physical Reward",  # Product Name
                reward.item_sku or f"REWARD-{reward.reward_type.value.upper()}",  # SKU
                reward.item_weight or 0.5,  # Weight
                reward.item_length or 15,  # Length
                reward.item_breadth or 10,  # Breadth
                reward.item_height or 5,  # Height
                "",  # AWB Code (empty - to be filled by Shiprocket)
                "",  # Tracking Number (empty - to be filled by Shiprocket)
                "",  # Courier Name (empty - to be filled by Shiprocket)
            ]

            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to BytesIO buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        logger.info(f"Exported {len(rewards)} pending shipments to Excel")

        return buffer

    async def import_shipment_tracking_from_excel(
        self, file: UploadFile
    ) -> BulkShipmentUpdateResponse:
        """
        Import tracking data from Excel and update rewards.

        Expected columns (case-insensitive, flexible matching):
        - Internal ID or Order Reference: To match rewards
        - AWB Code or Tracking Number: The tracking code
        - Courier Name or Courier Partner: Courier company (optional)

        Args:
            file: Excel file upload

        Returns:
            BulkShipmentUpdateResponse with summary

        Raises:
            ValueError: If file format invalid or required columns missing
        """
        # Validate file type
        if not file.filename.endswith((".xlsx", ".xls")):
            raise ValueError("File must be Excel format (.xlsx or .xls)")

        # Read file content
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active

        # Find header row and column mappings
        headers = {}
        for col_num, cell in enumerate(ws[1], 1):
            header = str(cell.value).lower().strip()
            headers[header] = col_num

        # Find required columns (flexible matching)
        internal_id_col = None
        order_ref_col = None
        awb_col = None
        courier_col = None

        for header, col_num in headers.items():
            if "internal" in header and "id" in header:
                internal_id_col = col_num
            elif "order" in header and ("reference" in header or "ref" in header or "id" in header):
                order_ref_col = col_num
            elif "awb" in header or ("tracking" in header and "number" in header):
                awb_col = col_num
            elif "courier" in header and ("name" in header or "partner" in header):
                courier_col = col_num

        # Validate required columns exist
        if not (internal_id_col or order_ref_col):
            raise ValueError("Excel must contain 'Internal ID' or 'Order Reference' column")
        if not awb_col:
            raise ValueError("Excel must contain 'AWB Code' or 'Tracking Number' column")

        # Process rows
        total_rows = 0
        successful_updates = 0
        failed_updates = 0
        errors = []

        for row_num, row in enumerate(ws.iter_rows(min_row=2), 2):
            total_rows += 1

            try:
                # Get values
                internal_id = row[internal_id_col - 1].value if internal_id_col else None
                order_ref = row[order_ref_col - 1].value if order_ref_col else None
                awb_code = row[awb_col - 1].value if awb_col else None
                courier_name = row[courier_col - 1].value if courier_col else None

                # Skip if no AWB code
                if not awb_code:
                    errors.append(f"Row {row_num}: Missing AWB code")
                    failed_updates += 1
                    continue

                # Find reward
                reward = None
                if internal_id:
                    reward = self.db.query(UserReward).filter(UserReward.id == internal_id).first()

                if not reward and order_ref:
                    # Extract reward ID from order reference (format: RNR-EVT-X-USR-Y-RWD-ZZZZZZZZ)
                    parts = order_ref.split("-RWD-")
                    if len(parts) == 2:
                        reward_id_prefix = parts[1][:8].lower()
                        # Find by ID prefix match
                        rewards = self.db.query(UserReward).filter(
                            UserReward.id.cast(str).ilike(f"{reward_id_prefix}%")
                        ).all()
                        if len(rewards) == 1:
                            reward = rewards[0]

                if not reward:
                    errors.append(f"Row {row_num}: Reward not found (ID: {internal_id or order_ref})")
                    failed_updates += 1
                    continue

                # Update reward
                reward.tracking_number = str(awb_code).strip()
                if courier_name:
                    reward.courier_partner = str(courier_name).strip()
                reward.status = RewardStatus.SHIPPED.value
                reward.shipped_at = datetime.utcnow()

                # Update status history
                if not reward.status_history:
                    reward.status_history = []
                history = reward.status_history if isinstance(reward.status_history, list) else []
                history.append({
                    "status": "shipped",
                    "timestamp": datetime.utcnow().isoformat(),
                    "source": "bulk_import",
                    "tracking_number": reward.tracking_number,
                    "courier": reward.courier_partner or "Unknown",
                })
                reward.status_history = history

                successful_updates += 1
                logger.info(f"Updated reward {reward.id} with AWB {awb_code}")

            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")
                failed_updates += 1
                logger.error(f"Failed to process row {row_num}: {str(e)}")

        # Commit all changes
        self.db.commit()

        logger.info(f"Bulk import complete: {successful_updates}/{total_rows} successful")

        return BulkShipmentUpdateResponse(
            total_rows=total_rows,
            successful_updates=successful_updates,
            failed_updates=failed_updates,
            errors=errors,
        )

    def toggle_shipping_verification(
        self, reward_id: str, verify: bool, admin_id: int
    ) -> UserReward:
        """
        Toggle shipping verification status (de-verify only).

        This method only handles de-verification. Verification should be done
        through AdminVerifyShippingModal which handles verification + unlock atomically.

        De-verification:
        - Sets is_verified = False
        - Hides tracking from user (tracking_visible_to_user = False)
        - Preserves all tracking data
        - Keeps reward record intact

        Business Rules:
        - Cannot de-verify if status = shipped or delivered
        - Preserves manual_tracking_url for re-verification

        Args:
            reward_id: UUID of reward
            verify: Must be False (True redirects to modal)
            admin_id: Admin user ID performing the action

        Returns:
            Updated UserReward

        Raises:
            NotFoundException: Reward not found
            ValueError: Trying to de-verify shipped/delivered reward
        """
        # Get reward
        reward = self.db.query(UserReward).filter(UserReward.id == reward_id).first()

        if not reward:
            raise NotFoundException("Reward", reward_id)

        # Business rule: cannot de-verify shipped/delivered rewards
        if not verify and reward.status in [RewardStatus.TRACKING_ORDER.value, RewardStatus.DELIVERED.value]:
            raise ValueError(
                f"Cannot de-verify reward with status '{reward.status}'. "
                "Reward has already been shipped or delivered."
            )

        # De-verification logic
        if not verify:
            reward.is_verified = False
            reward.tracking_visible_to_user = False
            reward.verified_at = None
            reward.verified_by_admin_id = None

            # Update status history
            if not reward.status_history:
                reward.status_history = []
            history = reward.status_history if isinstance(reward.status_history, list) else []
            history.append({
                "action": "de_verified",
                "timestamp": datetime.utcnow().isoformat(),
                "admin_id": admin_id,
                "note": "Shipping details de-verified by admin",
            })
            reward.status_history = history

            logger.info(f"Reward {reward_id} de-verified by admin {admin_id}")

        self.db.commit()
        self.db.refresh(reward)

        return reward

    def update_tracking_url(
        self, reward_id: str, tracking_url: str | None, admin_id: int
    ) -> UserReward:
        """
        Add, update, or delete tracking URL inline.

        Operations:
        - Add/Update: tracking_url = "https://..."
        - Delete: tracking_url = None

        Business Rules:
        - Reward must be verified (is_verified = True)
        - Auto-enable tracking_visible_to_user when URL added
        - Auto-disable tracking_visible_to_user when URL deleted
        - Updates manual_tracking_url field

        Args:
            reward_id: UUID of reward
            tracking_url: Tracking URL or None to delete
            admin_id: Admin user ID performing the action

        Returns:
            Updated UserReward

        Raises:
            NotFoundException: Reward not found
            ValueError: Reward not verified or invalid URL
        """
        # Get reward
        reward = self.db.query(UserReward).filter(UserReward.id == reward_id).first()

        if not reward:
            raise NotFoundException("Reward", reward_id)

        # Business rule: must be verified
        if not reward.is_verified:
            raise ValueError(
                "Cannot update tracking URL for unverified reward. "
                "Please verify shipping details first."
            )

        # Validate URL format (basic check)
        if tracking_url and not tracking_url.strip():
            raise ValueError("Tracking URL cannot be empty. Use null to delete.")

        # Update or delete tracking URL
        if tracking_url:
            # Add/Update
            reward.manual_tracking_url = tracking_url.strip()
            reward.tracking_visible_to_user = True  # Auto-enable visibility
            reward.tracking_imported_at = datetime.utcnow()
            reward.tracking_imported_by_admin_id = admin_id

            action = "tracking_url_updated" if reward.manual_tracking_url else "tracking_url_added"
            logger.info(f"Tracking URL {action} for reward {reward_id} by admin {admin_id}")
        else:
            # Delete
            reward.manual_tracking_url = None
            reward.tracking_visible_to_user = False  # Auto-disable visibility
            reward.tracking_imported_at = None
            reward.tracking_imported_by_admin_id = None

            action = "tracking_url_deleted"
            logger.info(f"Tracking URL deleted for reward {reward_id} by admin {admin_id}")

        # Update status history
        if not reward.status_history:
            reward.status_history = []
        history = reward.status_history if isinstance(reward.status_history, list) else []
        history.append({
            "action": action,
            "timestamp": datetime.utcnow().isoformat(),
            "admin_id": admin_id,
            "tracking_url": tracking_url if tracking_url else None,
        })
        reward.status_history = history

        self.db.commit()
        self.db.refresh(reward)

        return reward
