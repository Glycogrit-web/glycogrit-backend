"""
CSV Processing Service for Bulk Certificate Upload
Processes Autocrat-generated CSV/XLSX with certificate URLs from Google Drive
"""
import csv
import io
import logging
from datetime import datetime
from typing import Dict, List, Tuple

import openpyxl
from sqlalchemy.orm import Session

from app.core.exceptions import ValidationException
from app.models.user import User
from app.modules.certificates.config import certificate_config
from app.modules.registrations.domain.registration import Registration
from app.services.base import BaseService

logger = logging.getLogger(__name__)


class CSVProcessorService(BaseService):
    """Service for processing certificate CSV/XLSX uploads from Autocrat"""

    def __init__(self, db: Session):
        super().__init__(db)

    def convert_xlsx_to_csv(self, xlsx_content: bytes, sheet_name: str | None = None) -> str:
        """
        Convert XLSX file to CSV format

        Args:
            xlsx_content: XLSX file bytes
            sheet_name: Optional sheet name (uses first sheet if not provided)

        Returns:
            CSV content as string

        Raises:
            ValidationException: If XLSX is invalid or sheet not found
        """
        try:
            # Load workbook from bytes
            workbook = openpyxl.load_workbook(io.BytesIO(xlsx_content), data_only=True)

            # Select sheet
            if sheet_name:
                # Use specified sheet name
                if sheet_name not in workbook.sheetnames:
                    available = "', '".join(workbook.sheetnames)
                    raise ValidationException(
                        f"Sheet '{sheet_name}' not found. Available sheets: '{available}'"
                    )
                sheet = workbook[sheet_name]
                logger.info(f"📊 Using specified sheet: '{sheet_name}'")
            else:
                # Use first sheet
                sheet = workbook.active
                logger.info(f"📊 Using first sheet: '{sheet.title}' (from {len(workbook.sheetnames)} total sheets)")

            # Convert to CSV
            output = io.StringIO()
            writer = csv.writer(output)

            for row in sheet.iter_rows(values_only=True):
                # Skip completely empty rows
                if all(cell is None or cell == '' for cell in row):
                    continue

                # Convert cells to strings, handling None values
                row_data = [str(cell) if cell is not None else '' for cell in row]
                writer.writerow(row_data)

            csv_content = output.getvalue()
            output.close()

            logger.info(f"✅ Converted XLSX to CSV ({len(csv_content)} bytes)")
            return csv_content

        except ValidationException:
            # Re-raise validation exceptions
            raise
        except Exception as e:
            logger.error(f"❌ Failed to convert XLSX to CSV: {e}")
            raise ValidationException(f"Invalid Excel file: {str(e)}")

    def validate_csv_format(self, csv_content: str) -> Tuple[bool, str, List[str]]:
        """
        Validate CSV has required columns with flexible pattern matching for Autocrat

        Returns:
            Tuple of (is_valid, error_message, column_names)
        """
        try:
            csv_file = io.StringIO(csv_content)
            reader = csv.DictReader(csv_file)

            if not reader.fieldnames:
                return False, "CSV file is empty or has no headers", []

            columns = [col.strip() for col in reader.fieldnames]

            # Check for required columns (case-insensitive)
            columns_lower = [col.lower() for col in columns]

            # Check for email column (exact match)
            if certificate_config.REQUIRED_EMAIL_COLUMN.lower() not in columns_lower:
                return False, f"Missing required column: {certificate_config.REQUIRED_EMAIL_COLUMN}", columns

            # Check for certificate URL column (pattern match for Autocrat flexibility)
            cert_url_found = False
            for col_lower in columns_lower:
                for pattern in certificate_config.CERTIFICATE_URL_PATTERNS:
                    if pattern in col_lower:
                        cert_url_found = True
                        break
                if cert_url_found:
                    break

            if not cert_url_found:
                patterns_display = "', '".join(certificate_config.CERTIFICATE_URL_PATTERNS)
                return False, f"Missing certificate URL column. Expected column containing: '{patterns_display}'", columns

            return True, "", columns

        except Exception as e:
            return False, f"Invalid CSV format: {str(e)}", []

    def process_certificate_csv(
        self,
        csv_content: str,
        event_id: int,
        uploaded_by_admin_id: int
    ) -> Dict:
        """
        Process CSV and update registrations with certificate URLs

        Args:
            csv_content: CSV file content as string
            event_id: Event ID to match registrations
            uploaded_by_admin_id: Admin user ID uploading the CSV

        Returns:
            Dictionary with processing statistics
        """
        # Validate format
        is_valid, error, columns = self.validate_csv_format(csv_content)
        if not is_valid:
            raise ValidationException(error)

        # Find column indices (case-insensitive with flexible pattern matching for Autocrat)
        columns_lower = {col.lower(): i for i, col in enumerate(columns)}

        # Find email column (exact match)
        email_idx = columns_lower.get(certificate_config.REQUIRED_EMAIL_COLUMN.lower())

        # Find certificate URL column (pattern match for Autocrat-generated names)
        # Examples: "Merged Doc URL - Auto Certificate", "Link to merged Doc - Auto Certificate"
        # Strategy: Find ALL matching columns, prefer more specific patterns first
        cert_url_idx = None
        best_pattern_priority = len(certificate_config.CERTIFICATE_URL_PATTERNS)  # Lower is better

        for col_lower, idx in columns_lower.items():
            for priority, pattern in enumerate(certificate_config.CERTIFICATE_URL_PATTERNS):
                if pattern in col_lower:
                    # Prefer columns with higher priority patterns (earlier in list)
                    if priority < best_pattern_priority:
                        cert_url_idx = idx
                        best_pattern_priority = priority
                        logger.info(f"📋 Found certificate URL column: '{columns[idx]}' (matched pattern: '{pattern}', priority: {priority})")
                    break

        # Find distance column (optional, case-insensitive)
        distance_idx = columns_lower.get('distance')

        # Find activity type/sport column (optional, multiple variations)
        activity_type_idx = (
            columns_lower.get('sport') or
            columns_lower.get('activity_type') or
            columns_lower.get('activity type')
        )

        if email_idx is None or cert_url_idx is None:
            raise ValidationException("Could not find required columns in CSV. Ensure 'email' and a certificate URL column exist.")

        # Process rows
        csv_file = io.StringIO(csv_content)
        reader = csv.reader(csv_file)
        next(reader)  # Skip header

        stats = {
            'total_rows': 0,
            'successful': 0,
            'failed': 0,
            'not_found': 0,
            'already_has_certificate': 0,
            'errors': []
        }

        for row_num, row in enumerate(reader, start=2):
            stats['total_rows'] += 1

            try:
                # Validate row has enough columns for required fields
                required_indices = [email_idx, cert_url_idx]
                if len(row) <= max(idx for idx in required_indices if idx is not None):
                    stats['failed'] += 1
                    stats['errors'].append(f"Row {row_num}: Incomplete data")
                    continue

                email = row[email_idx].strip().lower()
                cert_url = row[cert_url_idx].strip() if cert_url_idx < len(row) else ""

                # Extract optional distance field (case-insensitive)
                distance_str = ""
                distance_value = None
                if distance_idx is not None and distance_idx < len(row):
                    distance_str = row[distance_idx].strip()
                    if distance_str:
                        try:
                            distance_value = float(distance_str)
                            if distance_value < 0:
                                logger.warning(f"Row {row_num}: Negative distance {distance_value}, setting to 0")
                                distance_value = 0
                        except ValueError:
                            logger.warning(f"Row {row_num}: Invalid distance value '{distance_str}', ignoring")
                            distance_value = None

                # Extract optional activity type/sport field (case-insensitive)
                activity_type_value = None
                if activity_type_idx is not None and activity_type_idx < len(row):
                    activity_type_str = row[activity_type_idx].strip().lower()
                    if activity_type_str:
                        # Normalize common variations
                        activity_type_map = {
                            'run': 'running', 'running': 'running', 'runner': 'running',
                            'cycle': 'cycling', 'cycling': 'cycling', 'bike': 'cycling', 'biking': 'cycling',
                            'walk': 'walking', 'walking': 'walking'
                        }
                        activity_type_value = activity_type_map.get(activity_type_str, activity_type_str)

                # SECURITY: Handle null/empty values explicitly
                # If email is empty, skip row
                if not email:
                    stats['failed'] += 1
                    stats['errors'].append(f"Row {row_num}: Empty email")
                    continue

                # IMPORTANT: Allow empty/null URLs to clear certificate URLs
                # This enables admin to remove certificates via CSV import
                # Empty URL will set external_certificate_url to empty string
                if not cert_url:
                    logger.warning(f"Row {row_num}: Empty URL for {email} - will clear certificate URL if exists")
                    # Continue processing to allow URL clearing

                # Find user by email
                user = self.db.query(User).filter(
                    User.email == email
                ).first()

                if not user:
                    stats['not_found'] += 1
                    stats['errors'].append(f"Row {row_num}: User not found - {email}")
                    continue

                # Find registration for this event
                registration = self.db.query(Registration).filter(
                    Registration.user_id == user.id,
                    Registration.event_id == event_id
                ).first()

                if not registration:
                    stats['not_found'] += 1
                    stats['errors'].append(
                        f"Row {row_num}: No registration found - {email} for event {event_id}"
                    )
                    continue

                # SECURITY: Track if overwriting existing certificate
                had_existing_cert = bool(registration.external_certificate_url)
                if had_existing_cert:
                    stats['already_has_certificate'] += 1
                    if cert_url:
                        logger.info(
                            f"Row {row_num}: Registration {registration.id} REPLACING certificate URL"
                        )
                    else:
                        logger.warning(
                            f"Row {row_num}: Registration {registration.id} CLEARING certificate URL (empty value provided)"
                        )

                # IMPORTANT: Update registration with certificate URL
                # This REPLACES any existing URL (even if new URL is empty/null)
                # Empty URLs will clear the certificate, allowing admin to revoke access
                registration.external_certificate_url = cert_url if cert_url else None

                # Update distance from CSV if provided
                registration.external_certificate_distance = distance_value

                # Update activity type from CSV if provided
                registration.external_certificate_activity_type = activity_type_value

                # SECURITY: Always lock certificate when updating URL (even if clearing)
                # Admin must explicitly unlock after URL change
                registration.external_certificate_unlocked = False

                # Update metadata
                registration.external_certificate_uploaded_at = datetime.utcnow()
                registration.external_certificate_uploaded_by = uploaded_by_admin_id

                stats['successful'] += 1

                if cert_url:
                    distance_info = f" (distance: {distance_value}km)" if distance_value else ""
                    sport_info = f", sport: {activity_type_value}" if activity_type_value else ""
                    logger.info(f"✅ Updated registration {registration.id} with certificate URL{distance_info}{sport_info} (locked)")
                else:
                    logger.info(f"✅ Cleared certificate URL for registration {registration.id}")

            except Exception as e:
                stats['failed'] += 1
                error_msg = f"Row {row_num}: {str(e)}"
                stats['errors'].append(error_msg)
                logger.error(error_msg)

        # Commit all changes
        try:
            self.db.commit()
            logger.info(
                f"📊 CSV Processing Complete: {stats['successful']}/{stats['total_rows']} successful"
            )
        except Exception as e:
            self.db.rollback()
            logger.error(f"❌ Failed to commit changes: {e}")
            raise ValidationException(f"Database commit failed: {str(e)}")

        return stats

    def get_registrations_with_certificates(self, event_id: int) -> List[Dict]:
        """
        Get all registrations that have external certificates for an event

        Args:
            event_id: Event ID

        Returns:
            List of registration dictionaries with certificate info
        """
        registrations = (
            self.db.query(Registration)
            .filter(
                Registration.event_id == event_id,
                Registration.external_certificate_url.isnot(None)
            )
            .all()
        )

        results = []
        for reg in registrations:
            results.append({
                'id': reg.id,
                'registration_number': reg.registration_number,
                'participant_name': reg.participant_name,
                'user_email': reg.user.email if reg.user else None,
                'external_certificate_url': reg.external_certificate_url,
                'external_certificate_unlocked': reg.external_certificate_unlocked,
                'external_certificate_uploaded_at': reg.external_certificate_uploaded_at,
            })

        return results
